
import random
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os
from datetime import datetime
import subprocess
import time
import sys
import argparse
import shutil
import glob
import json

def gen_xl_csh(json_file_name,no_of_cores,pathsel_widths,delaysel_widths,funclk_freq,random_sel_values):
    xl_file_base = "regression"
    xl_file_date = datetime.now().strftime('_%d_%m_%y')
    xl_file_time = datetime.now().strftime('_%H_%M_%S')
    xl_file_extension = ".xlsx"
    xl_file_name = xl_file_base + xl_file_date + xl_file_time + xl_file_extension
    regression_dir_name = "regression" + xl_file_date + xl_file_time

    counter = 1
    if(os.path.exists(xl_file_name)):
        while os.path.exists(xl_file_name):
            xl_file_name = xl_file_base + xl_file_date + xl_file_time +"_"+ str(counter)+ xl_file_extension
            counter = counter + 1
    wb = Workbook()
    ws = wb.active
    ws.title = "Regression Status"
    #file=open("Regression.csh",'w')
    pathsel_values = []
    delaysel_values = []
    pathsel_min_value = 1
    delaysel_min_value = 0
    for i in range (no_of_cores):
        pathsel_max_value = (2 ** pathsel_widths[i]) - 1
        delaysel_max_value = (2 ** delaysel_widths[i]) -1
        pathsel_mid_value = int(pathsel_max_value/2)
        if(pathsel_mid_value == 1):
            pathsel_mid_value = pathsel_mid_value + 1
        delaysel_mid_value = int(delaysel_max_value/2)
        pathsel_value = [pathsel_min_value, pathsel_mid_value, pathsel_max_value]
        delaysel_value = [delaysel_min_value, delaysel_mid_value, delaysel_max_value]
        pathsel_values.append(pathsel_value)
        delaysel_values.append(delaysel_value)
        #print(pathsel_values,delaysel_values)
    #print(pathsel_values,delaysel_values)
    data=[]
    first_row=["Sl. No"]
    first_row.append("Test Category")
    first_row.append("SV Testcase")
    first_row.append("C Testcase")
    for i in range (no_of_cores):
        first_row.append("Pathsel core"+str(i))
        first_row.append("Delaysel core"+str(i))
    first_row.append("Funclk Frequency in GHz")
    first_row.append("PM_BUFF_DELAY_IN_PS")
    first_row.append("Random sel values")
    first_row.append("Status")
    first_row.append("Coments")
    first_row.append("Make command")
    data.append(first_row)
    sl_no_count=1
    try:
        with open(json_file_name, 'r') as json_file:
            test_cases = json.load(json_file)
    except FileNotFoundError:
        print(f"Error: The file '{json_file_name}' was not found.")
    except json.JSONDecodeError:
        print(f"Error: Failed to decode JSON from the file '{json_file_name}'.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

    for category, tests in test_cases.items():
        sv_testcases = tests["sv_testcase"]
        c_testcases = tests["c_testcase"]
        if category == "negative" or category == "exhaustive" or category == "sanity":
            # Handle the negative category differently
            for sv_test, c_test in zip(sv_testcases, c_testcases):
                #print(f"  SV Testcase: {sv_test}, C Testcase: {c_test}")
                for i in range(1,2):
                    for j in range(1,2):
                        for k in range(1,2):
                            for l in range(1,2):
                                cur_row=[]
                                make_command="make -e sv_testcase=test_sarc_data_flow c_testcase=test_sarc_integrity"
                                cur_row.append(sl_no_count)
                                cur_row.append(category)
                                cur_row.append(sv_test)
                                cur_row.append(c_test)
                                for core_count in range (no_of_cores):
                                    cur_row.append(pathsel_values[core_count][k])
                                    cur_row.append(delaysel_values[core_count][l])
                                    make_command = make_command +" fixed_pathsel_core"+str(core_count)+"="+str((pathsel_values[core_count][k]))
                                    make_command = make_command +" fixed_delaysel_core"+str(core_count)+"="+str((delaysel_values[core_count][l]))
                                cur_row.append(j)
                                make_command = make_command +" funclk_freq="+str(j)
                                pm_buf_delay_in_ps = round(random.uniform(0.1,(1.0/j)),1)
                                cur_row.append(pm_buf_delay_in_ps)
                                make_command = make_command +" pm_buf_delay_in_ps="+str(pm_buf_delay_in_ps)+" all"
                                cur_row.append(i)
                                make_command = make_command +" random_sel_values="+str(i)
                                cur_row.append("NYR")
                                cur_row.append(None)
                                cur_row.append(make_command)
                                #file.write(make_command+"\n")
                                #print(make_command)
                                data.append(cur_row)
                                sl_no_count = sl_no_count + 1
        else:
            # Handle other categories with all possible combinations
            for sv_test in sv_testcases:
                for c_test in c_testcases:
                    for i in random_sel_values:
                        for j in funclk_freq:
                            for k in range(3):
                                for l in range(3):
                                    cur_row=[]
                                    make_command="make -e sv_testcase=test_sarc_data_flow c_testcase=test_sarc_integrity"
                                    cur_row.append(sl_no_count)
                                    cur_row.append(category)
                                    cur_row.append(sv_test)
                                    cur_row.append(c_test)
                                    for core_count in range (no_of_cores):
                                        cur_row.append(pathsel_values[core_count][k])
                                        cur_row.append(delaysel_values[core_count][l])
                                        make_command = make_command +" fixed_pathsel_core"+str(core_count)+"="+str((pathsel_values[core_count][k]))
                                        make_command = make_command +" fixed_delaysel_core"+str(core_count)+"="+str((delaysel_values[core_count][l]))
                                    cur_row.append(j)
                                    make_command = make_command +" funclk_freq="+str(j)
                                    pm_buf_delay_in_ps = round(random.uniform(0.1,(1.0/j)),1)
                                    cur_row.append(pm_buf_delay_in_ps)
                                    make_command = make_command +" pm_buf_delay_in_ps="+str(pm_buf_delay_in_ps)+" all"
                                    cur_row.append(i)
                                    make_command = make_command +" random_sel_values="+str(i)
                                    cur_row.append("NYR")
                                    cur_row.append(None)
                                    cur_row.append(make_command)
                                    #file.write(make_command+"\n")
                                    #print(make_command)
                                    data.append(cur_row)
                                    sl_no_count = sl_no_count + 1
    border_style = Border(
        left= Side(style='thin'),
        right= Side(style='thin'),
        top= Side(style='thin'),
        bottom= Side(style='thin')
    )
    for row_index,row in enumerate(data,start=1):
        for col_index,value in enumerate(row,start=1):
            cell=ws.cell(row=row_index,column=col_index, value=value)
            cell.border = border_style
        #ws.append(row)
    header_fill = PatternFill(start_color='FE9900',end_color='FE9900', fill_type='solid')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold = True)
    merge_column_index = "B"

    merge_start_row = None
    merge_previous_value = ws[f'{merge_column_index}{2}'].value

    for row in range (2, ws.max_row + 1):
        cell_value = ws[f'{merge_column_index}{row}'].value
        if(merge_previous_value == cell_value):
            continue
        else:
            if merge_start_row and merge_start_row < row -1:
                ws.merge_cells(f'{merge_column_index}{merge_start_row}:{merge_column_index}{row - 1}')
                ws[f'{merge_column_index}{merge_start_row}'].alignment = Alignment(horizontal='center',vertical='center')
            merge_start_row = row
            merge_previous_value = cell_value
    if merge_start_row and merge_start_row < ws.max_row:
        ws.merge_cells(f'{merge_column_index}{merge_start_row}:{merge_column_index}{ws.max_row}')
        ws[f'{merge_column_index}{merge_start_row}'].alignment = Alignment(horizontal='center',vertical='center')
    wb.save(xl_file_name)
    ws=auto_fit_coloumn_width(ws)
    wb.save(xl_file_name)
    print("Excel build done!!!!!!!")
    #file.close()
    return xl_file_name,regression_dir_name

def auto_fit_coloumn_width(ws):
    for coloumn in ws.columns:
        max_length=0
        column_letter = coloumn[0].column_letter
        for cell in coloumn:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjuste_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjuste_width
    return(ws)

def run_regression(no_of_cores,xl_file_name,regression_dir_name):
    file_name = xl_file_name
    wb = load_workbook(file_name)
    ws = wb.active
    command_col_index=10
    command_col_index = command_col_index + (no_of_cores * 2)
    status_col_index = command_col_index - 2
    column_data = []
    pass_fill = PatternFill(start_color='FF00B050',end_color='FF00B050', fill_type='solid')
    fail_fill = PatternFill(start_color='FFFF0000',end_color='FFFF0000', fill_type='solid')
    run_fill = PatternFill(start_color='00F0F9',end_color='00F0F9', fill_type='solid')
    err_fill = PatternFill(start_color='FFFFFF00',end_color='FFFFFF00', fill_type='solid')
    for row in range(2,ws.max_row + 1):
        ip_command = ws.cell(row=row,column=command_col_index).value
        if not ip_command:
            print(f"Skipping empty command at row {row}")
            continue
        print("Running: "+str(ip_command))
        ip_command = list(ip_command.split())
        ws.cell(row=row,column=status_col_index).fill =  run_fill
        ws.cell(row=row,column=status_col_index).value = "Running"
        wb.save(file_name)
        run_start_time = datetime.now()
        result = run_command(ip_command)
        if result is None:
            #print(f"run_command returned None for command: {ip_command}")
            result = [False, "try_again"]
        ws.cell(row=row,column=status_col_index).fill =  run_fill
        ws.cell(row=row,column=status_col_index).value = "Running"
        wb.save(file_name)
        reg_final_res=True
        flag = True
        run_count = 1
        issue = ''
        while (result[0] == False):
            #print(result)
            if(result[1]=="try_again"):
                result =list(run_command(ip_command))
                run_count = run_count +1
                if(run_count >= 10):
                    flag=False
                    #result[0] = True
                    print("Check bsub License")
                    ws.cell(row=row,column=status_col_index).value = "BSUB License issue"
                    ws.cell(row=row,column=status_col_index).fill =  err_fill
                    wb.save(file_name)
                    reg_final_res = False
                    issue = "Error in License"
            elif(result[1]=="terminate"):
                #print("Entered terminate")
                result[0] = True
                flag = False
                #ws.cell(row=row,column=status_col_index).value = "ERROR"
                #ws.cell(row=row,column=status_col_index).fill =  err_fill
                #wb.save(file_name)
                reg_final_res = False
                issue = "ERROR"
            else:
                print("Unexpected Error")
                result[0] = True
                reg_final_res = False
                flag = False
                issue = "Unexpected Error"
        if(flag == False):
            ws.cell(row=row,column=status_col_index).value = issue
            ws.cell(row=row,column=status_col_index).fill =  err_fill
            wb.save(file_name)
            break
        elif (result[1]=="PASS"):
            #print("Entered PASS")
            #result[0] = True
            #flag=True
            ws.cell(row=row,column=status_col_index).value = "PASS"
            ws.cell(row=row,column=status_col_index).fill =  pass_fill
            wb.save(file_name)
            move_current_result(run_start_time,regression_dir_name)
            reg_final_res = True
        elif (result[1]=="FAIL"):
            #print("Entered FAIL")
            #result[0] = True
            #flag=True
            ws.cell(row=row,column=status_col_index).value = "FAIL"
            ws.cell(row=row,column=status_col_index).fill =  fail_fill
            wb.save(file_name)
            move_current_result(run_start_time,regression_dir_name)
            reg_final_res = True
        else:
            print("Unexpected Error")
            break
        #print(result)
        print("Completed: "+' '.join(ip_command))
    wb.save(file_name)
    if(reg_final_res == True):
        print("Regression Successful")
    else:
        print("Regression Terminates")

def check_results(ip1,ip2):
    #ip1=str(ip1)
    flag = True
    if "Error" in ip2 or "No such file or directory" in ip2:
        flag = False
        file = open("compile.txt",'w')
        file.write(ip1)
        file.write(ip2)
        file.close()
    elif "UVM_ERROR" in ip1 or "UVM_FATAL" in ip1:
        err_start_index = ip1.find("UVM_ERROR :")
        err_end_index = ip1.find("\n",err_start_index)
        err_word = ip1[err_start_index:err_end_index]
        err_word = err_word.split()
        fatal_start_index = ip1.find("UVM_FATAL :")
        fatal_end_index = ip1.find("\n",fatal_start_index)
        fatal_word = ip1[fatal_start_index:fatal_end_index]
        fatal_word = fatal_word.split()
        if(int(err_word[2]) == 0) and (int(fatal_word[2]) == 0):
            flag = True
            print("Run successful")
            return [flag, "PASS"]
        else:
            flag = True
            print("Run successful")
            return [flag, "FAIL"]
        
    if(flag == False):
        if "[run_bsub] Error 255" in ip2:
            print("Job Not Submittted - trying again")
            return [flag, "try_again"]
        elif "[generate_force] Error 1" in ip2:
            print("Generation failed - check generate python file")
            #print(ip2)
            return [flag, "terminate"]
        elif "[compile] Error 2" in ip2:
            print("C Testcase compilation failed - check C files")
            #print(ip2)
            return [flag, "terminate"]
        elif "No such file or directory" in ip2:
            print("SV compilation failed - check SV files")
            #print(ip2)
            return [flag, "terminate"]
        elif "Error : Timeout" in ip2:
            print("Testcase is hanging or Give more time to run exceeding 1200 seconds")
            return [flag, "terminate"]
        else:
            print("Unexpected Error")
            #print(ip2)
            return [flag, "terminate"]

def run_command(command, timeout=600):
    try:
        result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=timeout)
        output1 = result.stdout.decode('utf-8')
        output2 = result.stderr.decode('utf-8')
        with open('log.txt', 'w') as file:
            file.write(output1)
            file.write(output2)
        return check_results(output1,output2)
    except subprocess.TimeoutExpired as e:
        print(f"Timeout expired: {e}")
        return check_results("Error : Timeout","0")
    except Exception as e:
        print(f"Unexpected error: {e}")
        return [False, "terminate"]

'''def run_command(cmd, timeout=600):
    process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

    start_time = time.time()
    end_time = start_time + timeout
    #file = open("test.txt",'w')
    print("entering run")
    while True:
        if process.poll() is not None:
            break
        #file.write(str(process.poll())+"\n")

        elapsed_time = time.time() - start_time
        remaining_time = end_time - time.time()

        percentage = min(int((elapsed_time / timeout) * 100), 100)

        sys.stdout.write(f"\rProgress: [{'#' * (percentage // 5)}{'.' * (20 - (percentage // 5))}] {percentage}% elapsed time: {elapsed_time:.2f} seconds")
        sys.stdout.flush()

        if remaining_time <= 0:
            process.terminate()
            sys.stdout.write("\nCommand timed out!\n")
            sys.stdout.flush()
            stdout, stderr = process.communicate()
            with open('log.txt', 'w') as file:
                file.write(stdout.decode('utf-8')+"\n\nERROR:\n\n")
                file.write(stderr.decode('utf-8'))
            file.close()
            return check_results("Error : Timeout","Error : Timeout")

        time.sleep(0.1)

    sys.stdout.write("\nCommand Completed\n")
    sys.stdout.flush()
    #file.close()
    print("exist run")

    stdout, stderr = process.communicate()
    print(stdout.decode('utf-8'))
    print(stderr.decode('utf-8'))

    return check_results(stdout.decode('utf-8'), stderr.decode('utf-8'))'''

def regression_analyze(no_of_cores,xl_file_name):
    wb = load_workbook(xl_file_name)
    ws = wb.active
    command_col_index=10
    command_col_index = command_col_index + (no_of_cores * 2)
    status_col_index = command_col_index - 2
    pass_count = 0
    fail_count = 0
    total_roows = 0
    border_style = Border(
        left= Side(style='thin'),
        right= Side(style='thin'),
        top= Side(style='thin'),
        bottom= Side(style='thin')
    )
    for row in range(2,ws.max_row + 1):
        total_roows = row
        status = ws.cell(row=row,column=status_col_index).value
        if(status == "PASS"):
            pass_count = pass_count + 1
        else:
            fail_count = fail_count + 1
    ws.cell(row=total_roows+2,column=3).value = "Total tests"
    ws.cell(row=total_roows+2,column=3).border = border_style
    #print(ws.cell(row=57,column=3).fill)
    ws.cell(row=total_roows+2,column=4).value = pass_count + fail_count
    ws.cell(row=total_roows+2,column=4).border = border_style

    ws.cell(row=total_roows+3,column=3).value = "Total pass"
    ws.cell(row=total_roows+3,column=3).border = border_style
    #print(ws.cell(row=58,column=3).fill)
    ws.cell(row=total_roows+3,column=4).value = pass_count 
    ws.cell(row=total_roows+3,column=4).border = border_style

    ws.cell(row=total_roows+4,column=3).value = "Total fail"
    ws.cell(row=total_roows+4,column=3).border = border_style
    #print(ws.cell(row=59,column=3).fill)
    ws.cell(row=total_roows+4,column=4).value = fail_count
    ws.cell(row=total_roows+4,column=4).border = border_style
    #print("Pass: "+str(pass_count)+" Fail: "+str(fail_count))
    wb.save(xl_file_name)

def move_current_result(start_time_str,regression_dir_name):
    current_time = datetime.now()# Convert to datetime objects
    end_time = current_time

    # Directory containing the log and fsdb files
    source_directory = '.'
    # Create a destination directory based on the start time of the regression
    destination_directory = regression_dir_name

    # Ensure the destination directory exists
    os.makedirs(destination_directory, exist_ok=True)

    # Patterns to match the files
    patterns = ["test_*.log", "test_*.fsdb"]

    # Filter and move the files
    for pattern in patterns:
        for file in glob.glob(os.path.join(source_directory, pattern)):
            filename = os.path.basename(file)
            file_mod_time = datetime.fromtimestamp(os.path.getmtime(file))
            # Debugging: Print the modification time of each file
            #print(f"File: {filename}, Modification Time: {file_mod_time}")

            if start_time_str <= file_mod_time <= end_time:
                destination_path = os.path.join(destination_directory, filename)
                shutil.move(file, destination_path)
                #print(f"Moved: {filename}")
    #print("File moving completed.")
